from datetime import datetime, timedelta
from airflow.utils.dates import days_ago
from airflow import DAG
from airflow.operators.python import PythonOperator
from pathlib import Path
from datetime import datetime
import psycopg2
from io import BytesIO
import openpyxl
import traceback
import collections 
# from  modules.my_excel_conf import excel_my_cnfg
# from  modules.conf import exportFile  

import io
import pandas as pd
import json 


def get_hello():
    path = Path(__file__).parent.parent.parent
    print (f'path of config file:  {path}')
    conf = json.load (open(path / ("config/email_conf.json")))  
    db_host = conf["db_host"]
    db_port = conf["db_port"]
    db_name = conf["db_name"]
    db_user = conf["db_user"]
    db_pass = conf["db_pass"]

    connection = psycopg2.connect(user=db_user,
                                                password=db_pass,
                                                host=db_host,
                                                port=db_port,
                                                database=db_name)
    cursor = connection.cursor()

    configs = {
        "errors":{
        "0001":{
            "en":"",
            "ru":"Неверное значение в ячейке, возможно есть ссылка, формула с ошибкой",
            "tj":"",
        },
        "0002":{
            "en":"",
            "ru":"Неверный тип данных в ячейке",
            "tj":"",
        },
        "0003":{
            "en":"",
            "ru":"Ячейке не должна быть пустой",
            "tj":"",
        },
        "0004":{
            "en":"",
            "ru":"Значение не должен быть отрицательной",
            "tj":"",
        },
        "0005":{
            "en":"",
            "ru":"В ячейке указанно слишком длинное значение",
            "tj":"",
        },
        "0006":{
            "en":"",
            "ru":"Значение не соответствует требуемому словарю",
            "tj":"",
        },
        "0007":{
            "en":"",
            "ru":"В файле найденно более 1000 ошибок, выполните тчательный анализ отчета и исправьте ошибки",
            "tj":"",
        },
        "0101":{
            "en":"",
            "ru":"Неправильное правило сравнения ячеек (тип возвращаемого значения",
            "tj":"",
        },
        "0201":{
            "en":"",
            "ru":"Срок сдачи отчета истек, обратитесь к контактному лицу НБТ по данному отчету",
            "tj":"" 
        },
        "0202":{
            "en":"",
            "ru":"Предоставление отчета невозможно, отчетный перид еще не закрыт",
            "tj":"" 
        },
        "0203":{
            "en":"",
            "ru":"Не найденна конфигурация по данному отчету",
            "tj":"" 
        },
        "0204":{
            "en":"",
            "ru":"Листы и их наименование в данном отчете не соответствует стандарту \n",
            "tj":"" 
        },
        "02041":{
            "en":"",
            "ru":"в файле присуствуют лишние листы \n",
            "tj":"" 
        },
        "02042":{
            "en":"",
            "ru":"в файле не хватает обязательных листов \n",
            "tj":"" 
        },
        "0205":{
            "en":"",
            "ru":"Для получения данного отчета не существует расписание \n",
            "tj":"" 
        },
    }
    } 
    emc = {
        "cells_codes":[
            ["A",1],
            ["B",2],
            ["C",3],
            ["D",4],
            ["E",5],
            ["F",6],
            ["G",7],
            ["H",8],
            ["I",9],
            ["J",10],
            ["K",11],
            ["L",12],
            ["M",13],
            ["N",14],
            ["O",15],
            ["P",16],
            ["Q",17],
            ["R",18],
            ["S",19],
            ["T",20],
            ["U",21],
            ["V",22],
            ["W",23],
            ["X",24],
            ["Y",25],
            ["Z",26],
            ["AA",27],
            ["AB",28],
            ["AC",29],
            ["AD",30],
            ["AE",31],
            ["AF",32],
            ["AG",33],
            ["AH",34],
            ["AI",35],
            ["AJ",36],
            ["AK",37],
            ["AL",38],
            ["AM",39],
            ["AN",40],
            ["AO",41],
            ["AP",42],
            ["AQ",43],
            ["AR",44],
            ["AS",45],
            ["AT",46],
            ["AU",47],
            ["AV",48],
            ["AW",49],
            ["AX",50],
            ["AY",51],
            ["AZ",52],
        ],
        "types":{
            'int':int,
            'str':str,
            'bool':bool,
            'object':object,
            'float':float,
        }
    }
    excelCells = emc["cells_codes"]
    types = emc["types"]
    def check_file_name(text):
        res = text.split('.')
        date = f'{res[3][-4:]}-{res[3][2:4]}-{res[3][0:2]}' 
        date = datetime.strptime(date, '%Y-%m-%d')
        print(date)
        return {
            "name":res[0],
            "version":res[1],
            "bic4":res[2],
            "date":str(date)
        }
    
    def selectOne(table,prop,value,get_arguments=None):
                    get_arguments = "id" if get_arguments is None else get_arguments
                    # print(get_arguments)
                    postgres_insert_query = f"select {get_arguments} from sma_stat_dep.{table} WHERE {value}" if prop is None else f"select {get_arguments} from sma_stat_dep.{table} WHERE {prop}='{value}'"
                    cursor.execute(postgres_insert_query)
                    # print(postgres_insert_query)
                    try:
                        if get_arguments=="id":
                            return cursor.fetchone()[0]
                        else:
                                return cursor.fetchone()
                    except:
                            return False
                    
    def selectAll(table,prop,value,get_arguments=None):
                    get_arguments = "id" if get_arguments is None else get_arguments
                    print(get_arguments)
                    postgres_insert_query = f"select {get_arguments} from sma_stat_dep.{table} WHERE {value}" if prop is None else f"select {get_arguments} from sma_stat_dep.{table} WHERE {prop}='{value}'"
                    cursor.execute(postgres_insert_query)
                    try:
                        if get_arguments=="id":
                            return tuple(map(lambda x:x[0],cursor.fetchall()))
                        else:
                                return cursor.fetchall()
                    except:
                            return False   
                                
    
    excelCells.reverse()
    def findeCEll(arg,arg2=None):
            if arg2 is None:
                res = arg
                # print(arg)
                for i in range(len(excelCells)):
                    if(arg.find(excelCells[i][0])>=0):
                        try:
                            res = [excelCells[i][1]-1,int(arg[len(excelCells[i][0]):])-1]
                        except:
                                return res
                return res
            else:
                if(type(arg)!=list):
                    arg = list(map(lambda x: int(x), arg.split(',')))
                else:
                    res = arg
                for i in excelCells:
                    if(int(arg[0])+1==i[1]):
                        res = f"{i[0]}{arg[1]+1}"  
                return res
            
    # print(findeCEll("0,112",True))
    # print(findeCEll("A112")) 

    def eval_formula(formula:str,wb):
        return eval(formula)
    
    def check_rule(rule,wb):
        result = eval_formula(rule["rule"],wb)
        # print(result)
        if type(result)!=bool:
            return {"error":True,"context":f"{configs['errors']['0101']['ru']} {type(result).__name__})"}
        if(result):
            return {"error":result,"context":rule["message"]}
        else:  
            return {"error":False,"context":""}


    def select_datas(file_upload_id):
        try:  
            postgres_insert_query = f"select * from sma_stat_dep.tbl_files WHERE id_file_upload={file_upload_id}"
            cursor.execute(postgres_insert_query) 
            mobile_records = cursor.fetchone() 
            file_id = mobile_records[0]
            logs = {
                        "count":0,
                        "comparisen_rules_count":0,
                        "context":'\n',
                        'upload_date':datetime.today().strftime('%Y-%m-%d %H:%M:%S')
                    }
            status = 1
            try:
                arg = check_file_name(mobile_records[3]) 
                report = selectOne('tbl_report_type','code',arg['name'],'id,report_period_type,submition_mode')
                print(report)
                bank_id = selectOne('tbl_entities','bic4',arg['bic4'])
                period_id = selectOne('tbl_period',None,f"type={report[1]} and to_date="+f"'{arg['date']}'")
                postgres_insert_query = f"select * from sma_stat_dep.tbl_schedule WHERE bank_id={bank_id} AND period_id={period_id} AND report_type_id={report[0]}"
                cursor.execute(postgres_insert_query)
                schedule_records = cursor.fetchone()
                to_date = selectOne('tbl_period',None,f"type={report[1]} and to_date="+f"'{arg['date']}'",'to_date')[0]

                postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET upload_status='2' WHERE id='{file_id}';"""
                cursor.execute(postgres_insert_query) 
                connection.commit()
                print(schedule_records[0])
                if(datetime.today().date()>to_date+timedelta(days=schedule_records[5])):
                    status = 5
                    logs["context"]=configs['errors']['0201']['ru']    
                    logs["count"]+=1
                    log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количество найденных ошибок --------------------{logs['count']}" 
                    postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                    cursor.execute(postgres_insert_query) 
                    connection.commit()
                    postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                    cursor.execute(postgres_insert_query1) 
                    connection.commit()
                    raise Exception(configs['errors']['0201']['ru'])
                elif(to_date>datetime.today().date()):
                    status = 5
                    logs["context"]+=configs['errors']['0201']['ru']    
                    logs["count"]+=1
                    log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количество найденных ошибок --------------------{logs['count']}" 
                    postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                    cursor.execute(postgres_insert_query) 
                    connection.commit()
                    postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                    cursor.execute(postgres_insert_query1) 
                    connection.commit()
                    raise Exception(configs['errors']['0202']['ru'])
                elif(schedule_records!=None): 
                    workbook_xml = BytesIO(bytes(mobile_records[4])) 
                    workbook_xml.seek(0)
                    # print(openpyxl.load_workbook(workbook_xml))

                    wb = openpyxl.load_workbook(workbook_xml)
                    # wb.close()

                    # wb.save(mobile_records[3])

                    # print(data[1][20:40],data[7][20:40])
                    # postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET upload_status='2' WHERE id='{file_id}';"""
                    # cursor.execute(postgres_insert_query) 
                    # connection.commit() 
                        # print(arg) 
                        
                        # print(data)
                    report_type = f"SELECT validation_config FROM sma_stat_dep.tbl_report_type WHERE code='{arg['name']}'" 
                    cursor.execute(report_type)
                    res = cursor.fetchone()
                    # print(cursor.fetchone())
                    if(res is None):
                            # print("Error while fetching data from PostgreSQL1", error,traceback.print_exc()) 
                            status = 5
                            logs["context"]+=configs['errors']['0203']['ru']    
                            logs["count"]+=1
                            log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количество найденных ошибок ---------------------------{(logs['count']+logs['comparisen_rules_count'])}" 
                            postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                            cursor.execute(postgres_insert_query) 
                            connection.commit()
                            postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                            cursor.execute(postgres_insert_query1) 
                            connection.commit()
                            raise Exception(configs['errors']['0203']['ru'])

                    cnf = json.loads(res[0])['tables']
                    # print(cnf)
                    configList = list(map(lambda x: x['sheet_name'], cnf))
                    mass = {"extra-sheets":[],"missing sheets":[]}
                    # print(wb.sheetnames)
                    for el in wb.sheetnames:
                            res = False
                            for e in configList:
                                if(el==e):
                                    res = True
                            if(res==False):
                                    mass["extra-sheets"].append(el)

                    for el in configList:
                            res = False
                            for e in wb.sheetnames:
                                if(el==e):
                                    res = True
                            if(res==False):
                                    mass["missing sheets"].append(el)
                    # print(wb.sheetnames)
                    # print(configList)

                        # ======================== password check===========================
                    for sheetname in wb.sheetnames:
                            print(wb[sheetname].protection)
                    #     if(wb[sheetname].protection.hashValue!='bRwS8h/cZaAnWYEWrHEun3Bp/C+19LCXaAcKANzfH7KNt9WrYl0slOu41YEj6R9dUC0OHktm892TBOYl+4yrVg=='):
                    #         status = 5
                    #         logs["context"]="Пароль файла не соответствует паролю стандартного шаблона отчета, используйте установленный стандартныйшаблонотчета"    
                    #         logs["count"]+=1
                    #         log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количествонайденныхошибок --------------------{logs['count']}" 
                    #         postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                    #         cursor.execute(postgres_insert_query) 
                    #         connection.commit()
                    #         postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                    #         cursor.execute(postgres_insert_query1) 
                    #         connection.commit()
                    #         raise Exception('Error password has changed',wb[sheetname].protection)        


                    if collections.Counter(wb.sheetnames) == collections.Counter(configList):
                            # entMass = {}
                            for table in cnf:
                                # entMass[table['sheet_name']] = {}
                                # if not have nodes
                                if('nodes' in table):
                                    logs["context"]+=f"Проверка и валидация ------------------------------------- листа {table['sheet_name']} \n"
                                    # data = pd.read_excel(mobile_records[3],index_col=None,header=None,sheet_name=table['sheet_name'], nrows=140, keep_default_na=False)

                                    # print(mobile_records[3])
                                    
                                    



                                    # df = pd.read_excel(mobile_records[3],index_col=None,header=None,sheet_name=table['sheet_name'], keep_default_na=False)
                                    # print ("Списки l1 и l2 одинаковые")
                                    cnfgMass = {}
                                    # print(table["table_type"])
                                        
                                    # print(cnfgMass.keys())
                                    # print(table['nodes'])
                                    # startTrConf = table['nodes'][0]['attribute'][0]['cell_address']
                                    table['nodes'] = sorted(table['nodes'],key=lambda x: (x.get('formula') is None, x.get('formula')),reverse=True)
                                    
                                    # print(schedule_records[0])
                                    # print(file_id)
                                    

                                    # print(table['nodes'])
                                    
                                    cnfList = set()

                                    if(table['table_type']=="fixed"):
                                        data = pd.read_excel(workbook_xml,engine='openpyxl',index_col=None,header=None,sheet_name=table['sheet_name'], keep_default_na=False)
                                        table_errors_count = 0
                                        for i in table['nodes']:
                                            # print(i)
                                            # print(i['cell_address'])
                                            cell = findeCEll(i['cell_address'])
                                            cnfgMass[f'{cell[0]},{cell[1]}'] = i 
                                            cnfList.add(i['code'])
                                        for k in data: 
                                            for i in range(len(data[k])):
                                                # print(f'{k},{i}')
                                                if (f'{k},{i}') in cnfgMass.keys():
                                                    excelCell = data[k][i]
                                                    configCell = cnfgMass[f'{k},{i}'] 
                                                    # print(configCell)
                                                    if(findeCEll(configCell['cell_address'])[0]!=k or findeCEll(configCell['cell_address'])[1]!=i):
                                                        print('error')
                                                    if((str(excelCell)!='nan' and type(excelCell)==str) and (configCell["is_empty_allowed"]==False and len(excelCell)==0)):  
                                                        logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0003']['ru']} {excelCell}\n"
                                                        logs["count"]+=1
                                                        continue
                                                    if(types[configCell["data_type"]]!=type(excelCell) and (str(excelCell)=='nan' and type(excelCell)!=str)): 
                                                        logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0001']['ru']} {excelCell}\n"
                                                        logs["count"]+=1
                                                        continue
                                                    if(types[configCell["data_type"]]!=type(excelCell) and excelCell!='' and (str(excelCell)!='nan')):
                                                        try:
                                                            type(int(excelCell)) 
                                                        except:  
                                                            logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0002']['ru']} {excelCell}\n"
                                                            logs["count"]+=1
                                                            continue
                                                    
                                                    if(len(str(excelCell))>configCell["length"]): 
                                                        logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0005']['ru']} {excelCell}\n"
                                                        logs["count"]+=1
                                                        print(f"{configCell['cell_address']}--------{configs['errors']['0005']['ru']}\n")
                                                        continue
                                                    if type(excelCell)==int:
                                                        if('is_negative_allowed' in configCell and (excelCell<0)!=configCell['is_negative_allowed']): 
                                                            logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0004']['ru']} {excelCell}\n"
                                                            logs["count"]+=1
                                                            continue
                                                    if((configCell.get('attr_allowed_value') is None)==False): 
                                                        if(len(configCell['attr_allowed_value'])>0):
                                                            if((excelCell in configCell['attr_allowed_value'])==False):
                                                                logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0006']['ru']} {excelCell}\n"
                                                                logs["count"]+=1
                                                                continue  
                                                    if((configCell.get('formula') is None)==False): 
                                                            eval_formula(configCell['formula'],wb)
                                                    
                                        if(logs["count"]<1):
                                            try:
                                                postgres_insert_query = f"""INSERT INTO sma_stat_dep.tbl_file_per_schedule (schedule_id,file_id)	VALUES ({schedule_records[0]},{file_id});"""
                                                cursor.execute(postgres_insert_query) 

                                            except (Exception, psycopg2.Error) as error: 
                                                print("Error while fetching data from PostgreSQL", error,traceback.print_exc()) 
                                            finally:
                                                connection.commit()

                                            db_query_values = []  
                                            tbl_file_per_schedule_id = selectOne('tbl_file_per_schedule',None,f"schedule_id={schedule_records[0]} and file_id="+f"'{file_id}'")
                                            ent = {}
                                            for el in selectAll('tbl_ent',None,f"code in {tuple(cnfList)}",'id,code'):
                                                ent.update({el[1]:el[0]})

                                            for i in table['nodes']: 
                                                cell = findeCEll(i['cell_address'])
                                                cnfgMass[f'{cell[0]},{cell[1]}'] = i 
                                            # print(cnfgMass['15,55'])
                                            # print(cnfgMass.keys())
                                            # print(ent)
                                            
                                            for k in data: 
                                                for i in range(len(data[k])):
                                                    obj = {}
                                                    if (f'{k},{i}') in cnfgMass.keys():
                                                        excelCell = data[k][i]
                                                        configCell = cnfgMass[f'{k},{i}']
                                                        if(configCell.get('comparison_rules')):
                                                            for rule in configCell.get('comparison_rules'):   
                                                                logs["context"]+=f"{configCell['cell_address']}--------{check_rule(rule,wb)['context']} {excelCell}\n"
                                                                logs["comparisen_rules_count"]+=1
                                                                continue
                                                        if(excelCell!=''):     
                                                            for attr in configCell['attribute']:
                                                                obj.update({attr['attr_type']:attr['attr_value']})
                                                            obj.update({'value':excelCell})
                                                            db_query_values.append([ent[configCell['code']],tbl_file_per_schedule_id,json.dumps(obj,ensure_ascii=False)])
                                                        # else:
                                                        #      print(configCell['cell_address'],excelCell)
                                            # print(db_query_values)

                                            # single line 
                                            # text = 'INSERT INTO sma_stat_dep.tbl_attr_values (ent_id,file_per_schedule_id,a_value) VALUES(%s, %s, %s)'
                                            # cursor.execute(text,mass[0]) 

                                            # multiple line
                                            execute_values(cursor,
                                            "INSERT INTO sma_stat_dep.tbl_attr_values (ent_id,file_per_schedule_id,a_value) VALUES %s",
                                            db_query_values)
                                            if(int(report[2])==1):    
                                                postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_schedule SET reporting_window='0' WHERE id='{schedule_records[0]}';"""
                                                cursor.execute(postgres_insert_query) 
                                                connection.commit()



                                    elif(table['table_type']=="unfixed"):
                                        print('read file ------------',datetime.today().strftime('%Y-%m-%d %H:%M:%S'))
                                        data = pd.read_excel(workbook_xml,engine='openpyxl',index_col=None,header=None,sheet_name=table['sheet_name'], keep_default_na=True).dropna(how='all')
                                        startTrConf = table['nodes'][0]['attribute'][0]['cell_address']
                                        table_errors_count = 0
                                        for node in table['nodes']:  
                                            # print(node["attribute"])
                                            for i in node["attribute"]:
                                                cell = findeCEll(i['cell_address'])
                                                cnfgMass[cell[0]] = i 
                                        print('start validation ------------',datetime.today().strftime('%Y-%m-%d %H:%M:%S'))
                                        for k in data:
                                            print(k) 
                                            conf_length_value = 1*10^cnfgMass[k]['length']-1
                                            for i in data.index:
                                                # if(table_errors_count<1000):    
                                                    if(k==findeCEll(cnfgMass[k]['cell_address'])[0]):
                                                        if (k) in cnfgMass.keys():
                                                            excelCell = data[k][i]
                                                            configCell = cnfgMass[k]
                                                            configCell['cell_address'] = findeCEll(f'{k},{i}',True)

                                                            

                                                            # if(findeCEll(configCell['cell_address'])[0]!=k or findeCEll(configCell['cell_address'])[1]!=i):
                                                            #     print('error')
                                                            # print(configCell["cell_address"],excelCell)
                                                            # print(excelCell)

                                                            cell_type = type(excelCell)
                                                            if(excelCell=='nan' and configCell["is_empty_allowed"]==False):  
                                                                logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0003']['ru']}\n"
                                                                table_errors_count+=1
                                                                continue
                                                            
                                                            if(types[configCell["data_type"]]!=cell_type):
                                                                try:
                                                                    type(int(excelCell)) 
                                                                except:  
                                                                    logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0002']['ru']}\n"
                                                                    table_errors_count+=1
                                                                    continue

                                                            if(cell_type==int):
                                                                if(conf_length_value<abs(excelCell)):
                                                                    logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0005']['ru']} {excelCell}\n"
                                                                    table_errors_count+=1
                                                                    continue

                                                            if(len(configCell['attr_allowed_value'])>0):
                                                                if((excelCell in configCell['attr_allowed_value'])==False):
                                                                        logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0006']['ru']}\n"
                                                                        table_errors_count+=1
                                                                        continue  
                                                                
                                                            # if(len(str(excelCell))>configCell["length"]): 
                                                            #     logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0005']['ru']} {excelCell}\n"
                                                            #     logs["count"]+=1
                                                            #     continue

                                                            # if type(excelCell)==int:
                                                            #     if(configCell['is_negative_allowed']==False and (excelCell<0)!=configCell['is_negative_allowed']): 
                                                            #         logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0004']['ru']} {excelCell}\n"
                                                            #         logs["count"]+=1
                                                            #         continue
                                                                

                                                            
                                                            # if((configCell.get('formula') is None)==False): 
                                                            #     eval_formula(configCell['formula'],wb)
                                                # else:
                                                #     logs["context"]+=f"{configCell['cell_address']}--------{configs['errors']['0007']['ru']}\n"
                                                #     logs["count"]+=1
                                                #     continue

                                        # print(logs)
                                        print('adding datas in to db ------------',datetime.today().strftime('%Y-%m-%d %H:%M:%S'))
                                        if(logs["count"]<1):
                                            try:
                                                postgres_insert_query = f"""INSERT INTO sma_stat_dep.tbl_file_per_schedule (schedule_id,file_id)	VALUES ({schedule_records[0]},{file_id});"""
                                                cursor.execute(postgres_insert_query) 

                                            except (Exception, psycopg2.Error) as error: 
                                                print("Error while fetching data from PostgreSQL", error,traceback.print_exc()) 
                                            finally:
                                                connection.commit()

                                            db_query_values = []
                                            cnfList = list(cnfgMass.values())
                                            tbl_file_per_schedule_id = selectOne('tbl_file_per_schedule',None,f"schedule_id={schedule_records[0]} and file_id="+f"'{file_id}'")
                                            ent = {}
                                            for el in selectAll('tbl_ent',None,f"code in {tuple(map(lambda att: att.get('attr_type'), cnfList))}",'id,code'):
                                                ent.update({el[1]:el[0]})


                                            for i in range(findeCEll(startTrConf)[1],len(data)):
                                                obj = {}
                                                for j in range(len(cnfList)): 
                                                    cell = findeCEll(cnfList[j]['cell_address'])
                                                    excelTr = data.iloc[i,:] 
                                                    if(pd.isna(excelTr[j])==False):
                                                        obj.update({cnfList[j]['attr_type']:excelTr[j]})
                                                    # else:
                                                    #      print(excelTr[j])

                                                    if((cnfList[j].get('comparison_rules') is None)==False):
                                                        for rule in cnfList[j].get('comparison_rules'):    
                                                            # print(cnfList[j]['cell_address'])
                                                            logs["context"]+=f"{cnfList[j]['cell_address']}--------{check_rule(rule,wb)['context']} {excelTr[j]}\n"
                                                            logs["comparisen_rules_count"]+=1
                                                            continue
                                                db_query_values.append([selectOne('tbl_ent','code',table['nodes'][0]['code']),tbl_file_per_schedule_id,json.dumps(obj,ensure_ascii=False,allow_nan=False)]) 
                                            # single line 
                                            # text = 'INSERT INTO sma_stat_dep.tbl_attr_values (ent_id,file_per_schedule_id,a_value) VALUES(%s, %s, %s)'
                                            # cursor.execute(text,mass[0]) 
                                            
                                            # print(db_query_values)
                                            # multiple line

                                            execute_values(cursor,
                                            "INSERT INTO sma_stat_dep.tbl_attr_values (ent_id,file_per_schedule_id,a_value) VALUES %s",
                                            db_query_values)
                                            if(int(report[2])==1):    
                                                postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_schedule SET reporting_window='0' WHERE id='{schedule_records[0]}';"""
                                                cursor.execute(postgres_insert_query) 
                                                connection.commit()

                                        print('end of validation------------',datetime.today().strftime('%Y-%m-%d %H:%M:%S'))
                                        
                                        if(table_errors_count>=1000):
                                            logs["context"]+=f"{configs['errors']['0007']['ru']}\n\n"
                                            table_errors_count+=1

                                        logs["count"]+=table_errors_count
                                    logs["context"]+=f"количество найденных ошибок на листе ---------------------------{(table_errors_count+logs['comparisen_rules_count'])}\n\n"

                                else:
                                    print('note nodes')
                                
                                if(logs['count']>0):
                                    status = 5
                                elif(logs['count']==0 and logs['comparisen_rules_count']>0):
                                    status = 3   
                                else:
                                    status = 4  
                    else:  
                            # print ("Списки l1 и l2 неодинаковые")
                            # print("Error ", error,traceback.print_exc()) 
                            status = 5
                            logs["context"]+=configs['errors']['0204']['ru']    
                            if(len(mass["extra-sheets"])>0):
                                    logs["count"]+=1
                                    text =''
                                    for el in mass["extra-sheets"]:
                                        text+= f'{el}, '
                                    logs["context"]+=f"{configs['errors']['02041']['ru']} {text}\n "
                            if(len(mass["missing sheets"])>0):
                                logs["count"]+=1
                                text =''
                                for el in mass["missing sheets"]:
                                        text+= f'{el}, '
                                logs["context"]+=f"{configs['errors']['02042']['ru']} {text}\n"
                            # print(logs)
                            logs["count"]+=1
                            print('adding logs in to db ------------',datetime.today().strftime('%Y-%m-%d %H:%M:%S'))

                            log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количество найденных ошибок---------------------------{(logs['count']+logs['comparisen_rules_count'])}" 
                            postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                            cursor.execute(postgres_insert_query) 
                            connection.commit()
                            postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                            cursor.execute(postgres_insert_query1) 
                            connection.commit()
                            raise Exception(configs['errors']['0204']['ru'])
                            
                    

                    # f = open("logs.txt", "w")
                    log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количеству найденных ошибок ---------------------------{(logs['count']+logs['comparisen_rules_count'])}"
                    # f.write(log_to_text)
                    # f.close() 

                    postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                    cursor.execute(postgres_insert_query) 
                    connection.commit()

                    postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                    cursor.execute(postgres_insert_query1) 
                    connection.commit()
                else:
                    status = 5
                    logs["context"]+=configs['errors']['0204']['ru']    
                    logs["count"]+=1
                    log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количество найденных ошибок ---------------------------{(logs['count']+logs['comparisen_rules_count'])}" 
                    postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                    cursor.execute(postgres_insert_query) 
                    connection.commit()
                    postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                    cursor.execute(postgres_insert_query1) 
                    connection.commit()
                    raise Exception(configs['errors']['0204']['ru'])
            except (Exception, psycopg2.Error) as error: 
                print("Error while fetching data from PostgreSQL", error,traceback.print_exc())
                status = 5
                # logs["context"]+="Для получения данного отчета не существует расписание"    
                # logs["count"]+=1
                log_to_text = f"Дата и время получения файла -------------------- {logs['upload_date']}\n\n {logs['context']}\n количество найденных ошибок ---------------------------{(logs['count']+logs['comparisen_rules_count'])}" 
                postgres_insert_query = f"""UPDATE sma_stat_dep.tbl_files SET logs='{log_to_text}', upload_status='{status}' WHERE id='{file_id}';"""
                cursor.execute(postgres_insert_query) 
                connection.commit()
                postgres_insert_query1 = f"""UPDATE sma_stat_dep.tbl_file_upload SET upload_status='{status}' WHERE id='{file_upload_id}';"""
                cursor.execute(postgres_insert_query1) 
                connection.commit()
                wb =None
            finally:
                if not wb is None: wb.close()

                
        except (Exception, psycopg2.Error) as error:
            print("Error while fetching data from PostgreSQL", error,traceback.print_exc())
    try:
        postgres_insert_query = f"select * from sma_stat_dep.tbl_files WHERE upload_status=1"
        # postgres_insert_query = f"select * from sma_stat_dep.tbl_file_upload WHERE id=32"
        cursor.execute(postgres_insert_query)
        mobile_records = cursor.fetchone()
        print('prinnnnnnnnnnnnnnnnt: ',mobile_records)
        # print(mobile_records[0])
        select_datas(mobile_records[1]) 
    except (Exception, psycopg2.Error) as error:
            print("Error while fetching data from PostgreSQL", error,traceback.print_exc())
    finally:
            if connection:
                cursor.close()
                connection.close()
                print("PostgreSQL connection is closed")


# def get_plugin():
#     from airflow_clickhouse_plugin.operators.clickhouse import ClickHouseOperator
#     print(f"clickhouseOperator!!!!!!! with version: {ClickHouseOperator.__class__}")
